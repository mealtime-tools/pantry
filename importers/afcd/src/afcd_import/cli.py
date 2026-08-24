"""Import selected AFCD Release 3 fields from caller-supplied workbooks."""

import hashlib
import json
import math
import os
import tempfile
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

import click
from mealtime_nutrients import kcal_from_kj
from openpyxl import load_workbook

DETAILS_SHEET = "Food details"
NUTRIENTS_SHEET = "All solids & liquids per 100 g"
RELEASE = "3"
DATA_PAGE = (
    "https://www.foodstandards.gov.au/science-data/food-nutrient-databases/"
    "afcd/data-files"
)
LICENCE_URL = (
    "https://www.foodstandards.gov.au/science-data/monitoringnutrients/"
    "afcd/datauserlicenceagreement"
)

FIELDS = {
    "key": "Public Food Key",
    "name": "Food Name",
    "kj": "Energy with dietary fibre, equated (kJ)",
    "protein": "Protein (g)",
    "fat": "Fat, total (g)",
    "fiber": "Total dietary fibre (g)",
    "sugar": "Total sugars (g)",
    "carbs": "Available carbohydrate, without sugar alcohols (g)",
}

NOTICE = f"""# Australian Food Composition Database notice

The `afcd.jsonl` shard is derived from the **Australian Food Composition
Database – Release 3**, © Food Standards Australia New Zealand (FSANZ).
Source and provenance: {DATA_PAGE}

The AFCD-derived shard is distributed under the FSANZ Data User Licence
Agreement, based on the Creative Commons Attribution-ShareAlike 3.0 Australia
licence: {LICENCE_URL}. This share-alike notice applies to the AFCD-derived
work, not to unrelated Pantry code. No FSANZ endorsement is implied, and no
FSANZ logo is used.

Changes made: Pantry selected the Public Food Key, food name, energy with
dietary fibre, protein, total fat, available carbohydrate without sugar
alcohols, total dietary fibre, and total sugars from the per-100 g sheet;
converted kilojoules to kilocalories by dividing by 4.184 and rounding to one
decimal place; represented the selected fields as deterministically sorted
JSONL; and added an empty brand field for Pantry's common record shape.

This work is based on Australian data. Australia data may not be appropriate
for use in other countries.

There are limitations associated with food composition databases. Food
composition data used in the database or databases may represent an average of
the nutrient content of a particular sample of foods and ingredients,
determined at a particular time. The nutrient composition of foods and
ingredients can vary substantially between batches and brands because of a
number of factors, including changes in season, processing practices and
ingredient source, and methods of calculation.
"""


def normalized_header(value: Any) -> str:
    """Make workbook line wrapping irrelevant while retaining exact words."""
    return " ".join(str(value or "").split())


def sha256(path: Path) -> str:
    """Compute the SHA-256 hex digest of a file."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def sheet(path: Path, name: str) -> tuple[Any, Any]:
    """Open and validate an AFCD Release 3 worksheet from a workbook."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    if name not in workbook.sheetnames:
        workbook.close()
        raise click.ClickException(f"{path.name} has no {name!r} sheet")
    selected = workbook[name]
    if "release 3" not in normalized_header(selected.cell(1, 1).value).lower():
        workbook.close()
        raise click.ClickException(
            f"{path.name} does not identify AFCD Release 3"
        )
    return workbook, selected


def columns(selected: Any, required: list[str]) -> dict[str, int]:
    """Map required column header names to their 1-based column indices."""
    headers: dict[str, int] = {}
    header_row = next(
        selected.iter_rows(min_row=3, max_row=3, values_only=True)
    )
    for index, value in enumerate(header_row, start=1):
        header = normalized_header(value)
        if header:
            if header in headers:
                raise click.ClickException(f"duplicate header {header!r}")
            headers[header] = index

    missing = [header for header in required if header not in headers]
    if missing:
        raise click.ClickException(f"missing required column: {missing[0]}")
    return {header: headers[header] for header in required}


def text_value(value: Any, row: int, label: str) -> str:
    """Return stripped text value or raise an exception if empty."""
    text = str(value or "").strip()
    if not text:
        raise click.ClickException(f"row {row} has no {label}")
    return text


def number_value(value: Any, row: int, label: str) -> float | int:
    """Validate and convert a nutrient cell value to a finite number."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise click.ClickException(
            f"row {row} has non-numeric {label}: {value!r}"
        )
    number = float(value)
    if not math.isfinite(number) or number < 0:
        raise click.ClickException(f"row {row} has invalid {label}: {value!r}")
    return int(number) if number.is_integer() else number


def shard_kcal(kilocalories: Decimal) -> float | int:
    """One energy figure as the shard writes it: a tenth, half-up, int if whole.

    Half-up rather than the banker's rounding `round` would apply, because the
    shipped shard was built that way and re-rounding must not move a published
    figure. This is formatting; the conversion is `kcal_from_kj`.
    """
    kcal = kilocalories.quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
    return int(kcal) if kcal == kcal.to_integral() else float(kcal)


def read_food_details(path: Path) -> dict[str, str]:
    """Read mapping of Public Food Key to Food Name from the details sheet."""
    workbook, selected = sheet(path, DETAILS_SHEET)
    try:
        wanted = [FIELDS["key"], FIELDS["name"]]
        found = columns(selected, wanted)
        foods: dict[str, str] = {}
        rows = selected.iter_rows(
            min_row=4, max_col=max(found.values()), values_only=True
        )
        for row_number, values in enumerate(rows, start=4):
            key_value = values[found[FIELDS["key"]] - 1]
            if key_value is None:
                continue
            key = text_value(key_value, row_number, "Public Food Key")
            name = text_value(
                values[found[FIELDS["name"]] - 1], row_number, "Food Name"
            )
            if key in foods:
                raise click.ClickException(f"duplicate Public Food Key: {key}")
            foods[key] = name
        return foods
    finally:
        workbook.close()


def nutrient_products(path: Path) -> dict[str, dict[str, Any]]:
    """Read product nutrient profiles per 100 g from the nutrients sheet."""
    workbook, selected = sheet(path, NUTRIENTS_SHEET)
    try:
        wanted = list(FIELDS.values())
        found = columns(selected, wanted)
        products: dict[str, dict[str, Any]] = {}
        rows = selected.iter_rows(
            min_row=4, max_col=max(found.values()), values_only=True
        )
        for row_number, row in enumerate(rows, start=4):
            key_value = row[found[FIELDS["key"]] - 1]
            if key_value is None:
                continue
            key = text_value(key_value, row_number, "Public Food Key")
            if key in products:
                raise click.ClickException(f"duplicate Public Food Key: {key}")

            values = {
                field: number_value(row[found[header] - 1], row_number, header)
                for field, header in FIELDS.items()
                if field not in {"key", "name"}
            }
            products[key] = {
                "id": key,
                "name": text_value(
                    row[found[FIELDS["name"]] - 1], row_number, "Food Name"
                ),
                "brand": "",
                # The sheet states kilojoules; a Pantry record holds kcal, so
                # it is converted here, formatted, and the kJ is not carried.
                "kcal": shard_kcal(kcal_from_kj(values["kj"])),
                "protein": values["protein"],
                "fat": values["fat"],
                "carbs": values["carbs"],
                "fiber": values["fiber"],
                "sugar": values["sugar"],
            }
        return products
    finally:
        workbook.close()


def atomic_write(path: Path, data: str) -> None:
    """Write data to destination path atomically using a temporary file."""
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary = tempfile.mkstemp(
        dir=path.parent, prefix=f".{path.name}."
    )
    try:
        with os.fdopen(
            descriptor, "w", encoding="utf-8", newline="\n"
        ) as stream:
            stream.write(data)
        os.chmod(temporary, 0o644)
        os.replace(temporary, path)
    except BaseException:
        Path(temporary).unlink(missing_ok=True)
        raise


@click.command()
@click.option(
    "--food-details",
    type=click.Path(path_type=Path, exists=True, dir_okay=False),
    required=True,
)
@click.option(
    "--nutrient-profiles",
    type=click.Path(path_type=Path, exists=True, dir_okay=False),
    required=True,
)
@click.option(
    "--out", type=click.Path(path_type=Path, file_okay=False), required=True
)
def main(food_details: Path, nutrient_profiles: Path, out: Path) -> None:
    """Import AFCD Release 3 workbooks supplied on the command line."""
    details = read_food_details(food_details)
    products = nutrient_products(nutrient_profiles)

    if details.keys() != products.keys():
        missing = sorted(details.keys() - products.keys())
        extra = sorted(products.keys() - details.keys())
        raise click.ClickException(
            "workbook Public Food Keys differ ("
            f"missing={missing[:1]}, extra={extra[:1]})"
        )
    for key, name in details.items():
        if products[key]["name"] != name:
            raise click.ClickException(f"Food Name mismatch for {key}")

    ordered = [products[key] for key in sorted(products)]
    jsonl = "".join(
        json.dumps(product, ensure_ascii=False, separators=(",", ":")) + "\n"
        for product in ordered
    )
    metadata = {
        "schema": 1,
        "source": "afcd",
        "release": "3",
        "record_count": len(ordered),
        "provenance": DATA_PAGE,
        "inputs": {
            "food_details": {"sha256": sha256(food_details)},
            "nutrient_profiles": {"sha256": sha256(nutrient_profiles)},
        },
        "sheet": NUTRIENTS_SHEET,
        "join_key": FIELDS["key"],
        "energy_conversion": (
            "kcal = energy_with_dietary_fibre_kj / 4.184; rounded to 0.1"
        ),
        "notice": "AFCD-NOTICE.md",
    }

    atomic_write(out / "afcd.jsonl", jsonl)
    atomic_write(
        out / "afcd.metadata.json",
        json.dumps(metadata, indent=2, ensure_ascii=False, sort_keys=True)
        + "\n",
    )
    atomic_write(out / "AFCD-NOTICE.md", NOTICE)
    click.echo(f"wrote {len(ordered)} AFCD products to {out / 'afcd.jsonl'}")


if __name__ == "__main__":
    main()
