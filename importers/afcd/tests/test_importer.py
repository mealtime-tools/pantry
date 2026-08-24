import hashlib
import json
from pathlib import Path

from afcd_import.cli import DETAILS_SHEET, FIELDS, NUTRIENTS_SHEET, main
from click.testing import CliRunner
from openpyxl import Workbook


def workbook(
    path: Path, *, mismatch: bool = False, target: bool = True
) -> None:
    book = Workbook()
    details = book.active
    details.title = DETAILS_SHEET
    details.cell(1, 1, "Australian Food Composition Database Release 3")
    details.cell(3, 1, FIELDS["key"])
    details.cell(3, 4, FIELDS["name"])
    details.cell(4, 1, "F000002")
    details.cell(4, 4, "Abalone")
    details.cell(5, 1, "F000001")
    details.cell(5, 4, "Water")
    book.save(path)

    if path.name.startswith("details"):
        return

    book = Workbook()
    wrong = book.active
    wrong.title = "Liquids only per 100 mL"
    wrong.cell(1, 1, "Australian Food Composition Database Release 3")
    wrong.cell(3, 1, FIELDS["key"])
    wrong.cell(4, 1, "WRONG")
    if target:
        nutrients = book.create_sheet(NUTRIENTS_SHEET)
        nutrients.cell(1, 1, "Australian Food Composition Database Release 3")
        columns = {
            FIELDS["key"]: 1,
            FIELDS["name"]: 4,
            FIELDS["kj"]: 5,
            FIELDS["protein"]: 8,
            FIELDS["fat"]: 10,
            FIELDS["fiber"]: 12,
            FIELDS["sugar"]: 20,
            FIELDS["carbs"]: 39,
        }
        for header, column in columns.items():
            nutrients.cell(3, column, header.replace(" ", "\n", 1))
        rows = [
            (
                "F000002",
                "Abalone" if not mismatch else "Other",
                418.4,
                20,
                2,
                0,
                0,
                1,
            ),
            ("F000001", "Water", 0, 0, 0, 0, 0, 0),
        ]
        for row, values in enumerate(rows, start=4):
            for value, field in zip(
                values,
                [
                    "key",
                    "name",
                    "kj",
                    "protein",
                    "fat",
                    "fiber",
                    "sugar",
                    "carbs",
                ],
            ):
                nutrients.cell(row, columns[FIELDS[field]], value)
    book.save(path)


def inputs(tmp_path: Path, *, mismatch: bool = False, target: bool = True):
    details = tmp_path / "details.xlsx"
    profiles = tmp_path / "profiles.xlsx"
    workbook(details)
    workbook(profiles, mismatch=mismatch, target=target)
    return details, profiles


def run_import(tmp_path: Path, **kwargs):
    details, profiles = inputs(tmp_path, **kwargs)
    out = tmp_path / "out"
    result = CliRunner().invoke(
        main,
        [
            "--food-details",
            str(details),
            "--nutrient-profiles",
            str(profiles),
            "--out",
            str(out),
        ],
    )
    return result, details, profiles, out


def test_imports_only_per_100g_sheet_and_maps_fields(tmp_path):
    result, _, _, out = run_import(tmp_path)
    assert result.exit_code == 0, result.output
    products = [
        json.loads(line)
        for line in (out / "afcd.jsonl").read_text().splitlines()
    ]
    assert [product["id"] for product in products] == ["F000001", "F000002"]
    assert products[0]["kcal"] == 0
    assert products[1] == {
        "id": "F000002",
        "name": "Abalone",
        "brand": "",
        "kcal": 100,
        "protein": 20,
        "fat": 2,
        "carbs": 1,
        "fiber": 0,
        "sugar": 0,
    }


def test_records_checksums_provenance_and_required_notice(tmp_path):
    result, details, profiles, out = run_import(tmp_path)
    assert result.exit_code == 0, result.output
    metadata = json.loads((out / "afcd.metadata.json").read_text())
    assert metadata["release"] == "3"
    assert metadata["record_count"] == 2
    assert (
        metadata["inputs"]["food_details"]["sha256"]
        == hashlib.sha256(details.read_bytes()).hexdigest()
    )
    assert (
        metadata["inputs"]["nutrient_profiles"]["sha256"]
        == hashlib.sha256(profiles.read_bytes()).hexdigest()
    )
    assert str(tmp_path) not in json.dumps(metadata)

    notice = (out / "AFCD-NOTICE.md").read_text()
    for phrase in [
        "Food Standards Australia New Zealand",
        "Attribution-ShareAlike 3.0 Australia",
        "Australia data may not be appropriate",
        "There are limitations associated",
        "No FSANZ endorsement",
        "converted kilojoules",
    ]:
        assert phrase in notice


def test_output_is_reproducible(tmp_path):
    result, details, profiles, out = run_import(tmp_path)
    assert result.exit_code == 0, result.output
    before = {path.name: path.read_bytes() for path in out.iterdir()}
    result = CliRunner().invoke(
        main,
        [
            "--food-details",
            str(details),
            "--nutrient-profiles",
            str(profiles),
            "--out",
            str(out),
        ],
    )
    assert result.exit_code == 0, result.output
    assert {path.name: path.read_bytes() for path in out.iterdir()} == before


def test_refuses_mismatched_food_names(tmp_path):
    result, _, _, out = run_import(tmp_path, mismatch=True)
    assert result.exit_code != 0
    assert "mismatch" in result.output.lower()
    assert not out.exists()


def test_requires_exact_per_100g_sheet(tmp_path):
    result, _, _, out = run_import(tmp_path, target=False)
    assert result.exit_code != 0
    assert NUTRIENTS_SHEET in result.output
    assert not out.exists()
